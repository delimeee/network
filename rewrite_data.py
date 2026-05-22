import openpyxl

def convert_excel_to_txt(excel_file_path, output_txt_path='bigboy.txt'):
    # Открываем Excel-файл
    workbook = openpyxl.load_workbook(excel_file_path, data_only=True)
    
    # Проверяем, что в файле есть нужные листы
    required_sheets = ["Потребители", "Подстанции (генерация)"]
    for sheet_name in required_sheets:
        if sheet_name not in workbook.sheetnames:
            print(f"Ошибка: Лист '{sheet_name}' не найден в Excel-файле!")
            return

    # Открываем файл для записи
    with open(output_txt_path, 'w', encoding='utf-8') as txt_file:
        
        # === ЧАСТЬ 1: ПОТРЕБИТЕЛИ ===
        sheet_consumers = workbook["Потребители"]
        for row in sheet_consumers.iter_rows(values_only=True):
            # Пропускаем пустые строки
            if all(cell is None for cell in row):
                continue
            
            # Проверяем, что в строке достаточно колонок (минимум 4, чтобы взять индексы 1, 2, 3)
            # Если колонок меньше, заполняем пустой строкой, чтобы скрипт не падал
            col2 = str(row[1]).strip() if len(row) > 1 and row[1] is not None else ""
            col3 = str(row[2]).strip() if len(row) > 2 and row[2] is not None else ""
            col4 = str(row[3]).strip() if len(row) > 3 and row[3] is not None else ""
            
            # Новый порядок: Координаты (3 и 4 столбцы), затем Потребление (2 столбец)
            clean_row = [col3, col4, col2]
            
            # Объединяем через пробел и добавляем в конец " 1"
            formatted_line = " ".join(clean_row) + " 1"
            txt_file.write(formatted_line + '\n')
            
        # === ЧАСТЬ 2: ПОДСТАНЦИИ ===
        sheet_substations = workbook["Подстанции (генерация)"]
        for row in sheet_substations.iter_rows(values_only=True):
            # Пропускаем пустые строки
            if all(cell is None for cell in row):
                continue
                
            # Извлекаем столбцы по новой логике
            col2 = str(row[1]).strip() if len(row) > 1 and row[1] is not None else ""
            col3 = str(row[2]).strip() if len(row) > 2 and row[2] is not None else ""
            col4 = str(row[3]).strip() if len(row) > 3 and row[3] is not None else ""
            
            # Новый порядок: Координаты (3 и 4 столбцы), затем Потребление (2 столбец)
            clean_row = [col3, col4, col2]
            
            # Объединяем через пробел и добавляем в конец " 0"
            formatted_line = " ".join(clean_row) + " 0"
            txt_file.write(formatted_line + '\n')

    print(f"Готово! Данные успешно переформатированы и сохранены в: {output_txt_path}")

# === ЗАПУСК СКРИПТА ===
# Укажите здесь правильное имя вашего Excel-файла
convert_excel_to_txt('Данные потребители и генерация.xlsx')